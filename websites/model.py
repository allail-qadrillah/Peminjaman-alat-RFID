import firebase_admin
from firebase_admin import credentials, firestore, db
import random
import string
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os
from dotenv import load_dotenv

load_dotenv()

credential = {
    "type": "service_account",
    "project_id": "web-peminjaman-alat-using-rfid",
    # "private_key_id": os.environ['PRIVATE_KEY_ID'],
    "private_key_id": "9181810463f3f13c24302fd8c88412286c555d1d",
    "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDYFLE0y6FM7frc\nwbiPPa/S2NNwXLDFn6JuZK3bDvr2BpU1hCbPGuQ9xOzGW8bacEAILFNeUX1Pregd\noFY1L/zt1b6zjeonOGw9N6/vNIksY54VXCQpqAENe0zFs5vJkfXnWZXSnHvaWuTt\nuQaektk6qHDct3mJTAtm+1MEa2OFEChcVMCkOcC1XsVtDi7I73EWS/QYvLSD6mJu\nfsyUlmK41VHfr71F6+7xjxpsd10g/41X/4nt5livWEvtMp96dZkzHbxVsRf1Qmyj\nXuOg8rsYMNiG15OQD9EWkuzOpezvucXfWgsLTxO3oXNhYCRF5E/X6/HrcjwOlhEY\nz0EXUw3TAgMBAAECggEAHr/rFov35gkOecXpAXFniQYpfhmitQI/okgaOaH1yCHY\nWqegwjjzIXCEZWyqwsErS5PLQBExV6/JksBozuQyW+zerjyphW/7xJvHzK8EP63k\nJulYOBlcH6gGnTKEc/idLubKcbtVIMDDR//cV5/8No3BP0JY6FdLCW+wjmDIkScg\nDqDunhRD9V02GiGTK4W10SLuUiRG1RcSs5e7xK43PYQ/lDHgl+x6pj4RaglTwpXD\n1Qx/XBYqXKuvp3o33eq2sBBmr3hs0EtACkuDfa3FTC0geXXcYEo0Nw+zmoXpmv8r\nQmx5qfxWWi0jO/XpRlKfKBtI6apYTnvd+924mu//AQKBgQD7yocHvk31zP8zYcfF\niErPqgM19SBlxFO4UyqQbB87xD9Nl/wgLwWVgBE4XocI9pEXWglw6KDqZQ8zVP9D\nkUgdVAYOFSQK+mzHv2KoadozcrAx08IxJGCote4C5pbDnemBwzEJHZHr266PPRed\nfrTLITmFrr9Udgcd4VN9GZW/gQKBgQDbsVolOKpJ6n3MKa3lNdJ/ePtS1EZ3NSDg\ncBSBpeGOcjK/sNZxemtX1zLYsQDGwl3QtRyBlfpYVdl16HTQ2YZdrvjAmdJezehd\nScdszSmO4qgh9ST65YY1QMcrmGWRXRCSiMZOnjyoK33+A7J5SPMlAYLhXTELIQWN\n32OPkEF3UwKBgQCw8hsCrhJLcszFZynk/2rTlT68ZI9n9RiAqLF43QL2FWwHM4/T\nbjgWX/G6E3QCMIS2TGfmGRU8o9iS0mCxMx/ivHGk4+Z4cJSRq0XB5OGa1jZMTexi\nbJYg+flRwIqi3g6DSpDjPwyVG8UIlH5MCAc12Rt7ftqmmGtX7Emn2JsMAQKBgES0\n4oxhfGBBhbYkEs5sVrinDXKKEL+XTgDjDIa/ahsye8yXcXlr38ZI9w3tzWBbyRtM\niO2+Cd02fBvz5xSp8uu5nPIfKTudCd6jBEtynlyhZhO4pygQQDjHWYK41orvoA8z\n11EfC7lMYMrgPSDRP7mDD3uLuZrRaPtFsc1kQBUhAoGBALlG0zaY3TwYTQv1f8Mc\nBsuGLymsM5wkc/oKXdVh01h9uf4cOmVJ6LY2aiC6poUoLqCumng74jzwGEOqDCHO\nbQaAWQs4knaHZpFoK4+jrfWF1y3VNoMRSz/x3OY2/e8utu4XQilW5wPPSvaPXVJi\nR7niTVubafP253/bPsjJ/K7K\n-----END PRIVATE KEY-----\n",
    "client_email": "firebase-adminsdk-xkeg9@web-peminjaman-alat-using-rfid.iam.gserviceaccount.com",
    "client_id": "107404317874574047447",
    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
    "token_uri": "https://oauth2.googleapis.com/token",
    "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
    "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-xkeg9%40web-peminjaman-alat-using-rfid.iam.gserviceaccount.com"
}
cred = credentials.Certificate(credential)
firebase_admin.initialize_app(cred, {
    'databaseURL': 'https://web-peminjaman-alat-using-rfid-default-rtdb.asia-southeast1.firebasedatabase.app'
})

firestore = firestore.client()


class util():
    def random_string(self, string_length=5):
        letters = string.ascii_letters
        return ''.join(random.choice(letters) for i in range(string_length))


class User(util):
    def __init__(self) -> None:
        pass

    def get_matakuliah_from_jurusan(self, id_kartu):
        list_matakuliah = []
        founded = False
        for mahasiswa in self.get_collection('mahasiswa'):
            if not founded:
                if mahasiswa.get('id_kartu') == id_kartu:
                    founded = True
                    for matakuliah in self.get_collection('matakuliah'):
                        if matakuliah.get('jurusan') == mahasiswa.get('nama_jurusan'):
                            list_matakuliah.append(matakuliah.get('nama'))
        return list_matakuliah
    
    def cek_proyektor(self, data_proyektor, id_proyektor, nama, nomor, kondisi):
        """mengembalikan True jika proyektor telah ada"""
        status = False
        for proyektor in data_proyektor:
            if proyektor.get('id_proyektor') == id_proyektor and proyektor.get('nama') == nama and proyektor.get('kondisi') == kondisi and proyektor.get('nomor') == nomor:
                status = True
                break
        return status

    def add_document(self, collection, document, data_document):
        return firestore.collection(collection).document(document).set(data_document)

    def get_collection(self, collection):
        data_collection = firestore.collection(collection).get()
        return [each.to_dict() for each in data_collection]

    def delete_document(self, collection, id):
        data_document = firestore.collection(
            collection).where('id', '==', id).get()
        return firestore.collection(collection).document(data_document[0].id).delete()

    def update_document(self, collection, document, data_update):
        return firestore.collection(collection).document(document).update(data_update)

    def change_status_proyektor(self, nama_proyektor):
        data_proyektor = self.get_collection('proyektor')
        for proyektor in data_proyektor:
            if proyektor['nama'] == nama_proyektor:
                self.update_document('proyektor', proyektor['id'], {
                                     'dipinjam': not proyektor['dipinjam']})

    def delete_peminjaman(self, proyektor):
        proyektor = firestore.collection('peminjaman').where(
            'proyektor', '==', proyektor).get()
        return firestore.collection('peminjaman').document(proyektor[0].to_dict()['id']).delete()

    def get_value_rtdb(self, key):
        return db.reference(key).get()

    def update_value_rtdb(self, key, value):
        ref = db.reference('/')
        ref.update({
            key: value
        })

    def find_mahasiswa(self):
        try:
            return [doc.to_dict() for doc in firestore.collection('mahasiswa').where(
                'id_kartu', '==', self.get_value_rtdb('id')).get()][0]
        except:
            return None

    def find_peminjaman(self):
        nama_mahasiswa = self.find_mahasiswa()['nama']
        list_peminjam = [doc.to_dict() for doc in firestore.collection('peminjaman').where(
            'nama', '==', nama_mahasiswa).get()]

        return list_peminjam

    def get_range_data(self, collection, start_time, end_time):
        data_range_waktu = []
        for item in self.get_collection(collection):
            if item['waktu'] >= start_time and item['waktu'] <= end_time:
                data_range_waktu.append(item)
        return data_range_waktu

    def export_to_excel(self, nama_file, start_time, end_time):

      wb = Workbook()
      ws = wb.active
      data_pengembalian = self.get_range_data(nama_file, start_time, end_time)

      ws.title = nama_file

      ws['A1'] = "No"
      ws['B1'] = f"Tanggal {nama_file}"
      ws['C1'] = "No Pinjam"
      ws['D1'] = "Mahasiswa"
      ws['E1'] = "Proyektor"
      ws['F1'] = "Perangkat Lainnya"
      ws['G1'] = "Mata Kuliah"
      ws['H1'] = "Ruang"
      ws['I1'] = "Dosen"

      for row, data in enumerate(data_pengembalian):

          perangkat_lainnya = " | ".join(
              [key for key, value in data['perangkat_lainnya'].items() if value])
          ws.append([row, data['waktu'], int(data['nomor_peminjaman']),
                    data['nama'], data['proyektor'], perangkat_lainnya,
                    data['matakuliah'], data['ruang'], data['dosen']])

      # Membuat objek PatternFill untuk mengatur warna
      fill = PatternFill(start_color='FFFF00',
                        end_color='FFFF00', fill_type='solid')

      # Mengiterasi semua sel pada baris dan mengatur warna fill
      for cell in ws[1]:
          cell.fill = fill

      from io import BytesIO
      file_stream = BytesIO()
      wb.save(file_stream)
      file_stream.seek(0)

      return file_stream
