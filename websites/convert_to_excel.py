from model import User
from openpyxl.styles import PatternFill
from openpyxl import Workbook

user = User()
# user.update_value_rtdb('id', '2')
start_time = "2023-03-15"  # Example start time
end_time = "2023-05-29"  # Example end time


def export_to_excel(nama_file, start_time, end_time):

  wb = Workbook()
  ws = wb.active
  data_pengembalian = user.get_range_data('peminjaman', start_time, end_time)

  ws.title = nama_file

  ws['A1'] = "No"
  ws['B1'] = "Tanggal Peminjaman"
  ws['C1'] = "No Pinjam"
  ws['D1'] = "Mahasiswa"
  ws['E1'] = "Proyektor"
  ws['F1'] = "Perangkat Lainnya"
  ws['G1'] = "Mata Kuliah"
  ws['H1'] = "Ruang"
  ws['I1'] = "Dosen"


  for row, data in enumerate(data_pengembalian):
      
      perangkat_lainnya = " | ".join(
          [key for key, value in data['perangkat_lainnya'].items() if value])
      ws.append([row, data['waktu'], int(data['nomor_peminjaman']), 
            data['nama'], data['proyektor'], perangkat_lainnya, 
            data['matakuliah'], data['ruang'], data['dosen']])

  # Membuat objek PatternFill untuk mengatur warna
  fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

  # Mengiterasi semua sel pada baris dan mengatur warna fill
  for cell in ws[1]:
      cell.fill = fill


  # print(data_pengembalian)
  wb.save(nama_file+".xlsx")


export_to_excel("coba", "2023-05-21T16:34", "2023-06-04T16:34")
